import re
import sqlite3
import pandas as pd
import streamlit as st
from io import BytesIO
from datetime import date
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def limpiar_nombre_archivo(nombre):
    nombre_limpio = re.sub(r"[^A-Za-z0-9_-]+", "_", str(nombre).strip())
    return nombre_limpio.strip("_") or "candidatos"


def asegurar_columnas_candidatos(conn):
    cursor = conn.cursor()
    cursor.execute("PRAGMA table_info(candidatos)")
    columnas = [fila[1] for fila in cursor.fetchall()]

    if "fecha_registro" not in columnas:
        cursor.execute("ALTER TABLE candidatos ADD COLUMN fecha_registro TEXT")

    if "estado_evaluacion" not in columnas:
        cursor.execute("ALTER TABLE candidatos ADD COLUMN estado_evaluacion TEXT DEFAULT 'Pendiente'")

    if "comentario_interno" not in columnas:
        cursor.execute("ALTER TABLE candidatos ADD COLUMN comentario_interno TEXT")

    cursor.execute("""
        UPDATE candidatos
        SET fecha_registro = ?
        WHERE fecha_registro IS NULL OR fecha_registro = ''
    """, (date.today().isoformat(),))

    cursor.execute("""
        UPDATE candidatos
        SET estado_evaluacion = 'Pendiente'
        WHERE estado_evaluacion IS NULL OR estado_evaluacion = ''
    """)

    conn.commit()


def crear_excel_candidatos_profesional(df, titulo):
    output = BytesIO()

    columnas = {
        "id": "ID",
        "fecha_registro": "Fecha de registro",
        "nombre_completo": "Nombre completo",
        "cargo_postula": "Cargo al que postula",
        "edad": "Edad",
        "estado_civil": "Estado civil",
        "hijos": "Hijos",
        "formacion_tercer_nivel": "Formacion tercer nivel",
        "titulo_cuarto_nivel": "Titulo cuarto nivel",
        "anos_experiencia": "Anos de experiencia",
        "empresa": "Ultima empresa",
        "cargo": "Ultimo cargo",
        "actividades": "Actividades realizadas",
        "sueldo_ultimo": "Ultimo sueldo",
        "aspiracion_salarial": "Aspiracion salarial",
        "motivo_salida": "Motivo de salida",
        "disponibilidad": "Disponibilidad",
        "estado_proceso": "Estado del proceso",
        "estado_evaluacion": "Estado de evaluacion",
        "comentario_interno": "Comentario interno",
        "observaciones": "Observaciones"
    }

    df_excel = df.rename(columns=columnas)
    orden = [nombre for nombre in columnas.values() if nombre in df_excel.columns]
    df_excel = df_excel[orden]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_excel.to_excel(writer, index=False, sheet_name="Candidatos", startrow=4)

        ws = writer.sheets["Candidatos"]
        max_col = max(len(df_excel.columns), 1)

        navy = "03045E"
        gold = "FDD817"
        white = "FFFFFF"
        light_blue = "EEF3FA"
        border_color = "B7C3D0"
        green = "1F8F4D"
        orange = "F59E0B"
        gray = "334155"

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)

        ws.cell(row=1, column=1).value = titulo
        ws.cell(row=2, column=1).value = "Reporte profesional de seleccion de personal"

        for row in [1, 2]:
            cell = ws.cell(row=row, column=1)
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(row=1, column=1).font = Font(bold=True, size=16, color=gold)
        ws.cell(row=2, column=1).font = Font(bold=True, size=12, color=white)

        thin_border = Border(
            left=Side(style="thin", color=border_color),
            right=Side(style="thin", color=border_color),
            top=Side(style="thin", color=border_color),
            bottom=Side(style="thin", color=border_color)
        )

        header_row = 5

        for cell in ws[header_row]:
            cell.font = Font(bold=True, color=white)
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        estado_columna = None
        for cell in ws[header_row]:
            if cell.value == "Estado de evaluacion":
                estado_columna = cell.column

        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, max_col=max_col):
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=light_blue)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

                if estado_columna and cell.column == estado_columna:
                    if cell.value == "Apto":
                        cell.fill = PatternFill("solid", fgColor=green)
                        cell.font = Font(bold=True, color=white)
                    elif cell.value == "No apto":
                        cell.fill = PatternFill("solid", fgColor=orange)
                        cell.font = Font(bold=True, color=navy)
                    else:
                        cell.fill = PatternFill("solid", fgColor=gray)
                        cell.font = Font(color=white)

        for column_cells in ws.iter_cols(min_col=1, max_col=max_col):
            column_letter = get_column_letter(column_cells[0].column)
            max_length = 12

            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[column_letter].width = min(max_length + 4, 42)

        ws.freeze_panes = "A6"
        ws.sheet_view.showGridLines = False

    output.seek(0)
    return output.getvalue()


def pintar_estado(valor):
    if valor == "Apto":
        return "background-color: #1f8f4d; color: white; font-weight: bold;"
    if valor == "No apto":
        return "background-color: #f59e0b; color: #03045E; font-weight: bold;"
    return "background-color: #334155; color: white;"


def aplicar_estilo_estado(df):
    styler = df.style

    if hasattr(styler, "map"):
        return styler.map(pintar_estado, subset=["estado_evaluacion"])

    return styler.applymap(pintar_estado, subset=["estado_evaluacion"])


def ver_candidatos():
    conn = sqlite3.connect("seleccion_personal.db", check_same_thread=False)

    st.title("Candidatos registrados")

    asegurar_columnas_candidatos(conn)

    df = pd.read_sql_query("""
        SELECT *
        FROM candidatos
        ORDER BY id DESC
    """, conn)

    if df.empty:
        st.info("Aun no hay candidatos registrados")
        conn.close()
        return

    df["fecha_registro"] = pd.to_datetime(df["fecha_registro"], errors="coerce")
    df["fecha_registro"] = df["fecha_registro"].fillna(pd.Timestamp(date.today()))
    df["estado_evaluacion"] = df["estado_evaluacion"].fillna("Pendiente")

    st.subheader("Filtros de busqueda")

    col1, col2, col3, col4, col5 = st.columns(5)

    with col1:
        busqueda = st.text_input("Buscar por nombre")

    columna_cargo = "cargo_postula" if "cargo_postula" in df.columns else "cargo"

    with col2:
        cargos = sorted(df[columna_cargo].dropna().astype(str).unique().tolist())
        cargo_filtro = st.selectbox("Filtrar por cargo", ["TODOS"] + cargos)

    with col3:
        fecha_inicio = st.date_input("Desde", value=df["fecha_registro"].min().date())

    with col4:
        fecha_fin = st.date_input("Hasta", value=df["fecha_registro"].max().date())

    with col5:
        estado_filtro = st.selectbox(
            "Estado",
            ["TODOS", "Pendiente", "Apto", "No apto"]
        )

    df_filtrado = df.copy()

    if busqueda.strip() and "nombre_completo" in df_filtrado.columns:
        df_filtrado = df_filtrado[
            df_filtrado["nombre_completo"]
            .fillna("")
            .str.contains(busqueda.strip(), case=False, na=False)
        ]

    if cargo_filtro != "TODOS":
        df_filtrado = df_filtrado[
            df_filtrado[columna_cargo].astype(str) == cargo_filtro
        ]

    if estado_filtro != "TODOS":
        df_filtrado = df_filtrado[
            df_filtrado["estado_evaluacion"].fillna("Pendiente") == estado_filtro
        ]

    df_filtrado = df_filtrado[
        (df_filtrado["fecha_registro"].dt.date >= fecha_inicio) &
        (df_filtrado["fecha_registro"].dt.date <= fecha_fin)
    ]

    st.caption(f"Resultados encontrados: {len(df_filtrado)}")

    columnas_vista = [
        "id",
        "fecha_registro",
        "nombre_completo",
        "cargo_postula",
        "edad",
        "estado_civil",
        "formacion_tercer_nivel",
        "anos_experiencia",
        "aspiracion_salarial",
        "disponibilidad",
        "estado_proceso",
        "estado_evaluacion",
        "comentario_interno",
        "observaciones"
    ]

    columnas_vista = [col for col in columnas_vista if col in df_filtrado.columns]

    df_vista = df_filtrado[columnas_vista].copy()
    df_vista["fecha_registro"] = df_vista["fecha_registro"].dt.strftime("%Y-%m-%d")

    if "estado_evaluacion" in df_vista.columns:
        st.dataframe(
            aplicar_estilo_estado(df_vista),
            use_container_width=True,
            hide_index=True
        )
    else:
        st.dataframe(
            df_vista,
            use_container_width=True,
            hide_index=True
        )

    st.divider()
    st.subheader("Descargar candidatos")

    tipo_descarga = st.radio(
        "Seleccione que desea descargar",
        ["Ultimos 3 candidatos", "Resultados filtrados", "Todos los candidatos"],
        horizontal=True
    )

    if tipo_descarga == "Ultimos 3 candidatos":
        df_descarga = df.sort_values("id", ascending=False).head(3)
        titulo_excel = "Ultimos 3 candidatos registrados"
        nombre_archivo = "ultimos_3_candidatos.xlsx"

    elif tipo_descarga == "Resultados filtrados":
        df_descarga = df_filtrado
        titulo_excel = "Candidatos filtrados"
        nombre_archivo = f"candidatos_filtrados_{limpiar_nombre_archivo(cargo_filtro)}.xlsx"

    else:
        df_descarga = df
        titulo_excel = "Todos los candidatos registrados"
        nombre_archivo = "todos_los_candidatos.xlsx"

    df_descarga = df_descarga.copy()
    df_descarga["fecha_registro"] = df_descarga["fecha_registro"].dt.strftime("%Y-%m-%d")

    excel = crear_excel_candidatos_profesional(df_descarga, titulo_excel)

    st.download_button(
        label="Descargar Excel profesional",
        data=excel,
        file_name=nombre_archivo,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    conn.close()
