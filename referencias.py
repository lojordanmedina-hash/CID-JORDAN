import re
import streamlit as st
import sqlite3
import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


NAVY = "03045E"
GOLD = "FDD817"
WHITE = "FFFFFF"
LIGHT_BLUE = "EEF3FA"
BORDER = "B7C3D0"


def limpiar_nombre_archivo(nombre):
    nombre_limpio = re.sub(r"[^A-Za-z0-9_-]+", "_", str(nombre).strip())
    return nombre_limpio.strip("_") or "candidato"


def preparar_referencias_para_excel(df):
    columnas_excel = {
        "empresa": "Empresa",
        "nombre_referente": "Nombre del referente",
        "cargo_referente": "Cargo del referente",
        "telefono": "Teléfono",
        "relacion": "Relacion",
        "desempeno": "Desempeno del candidato",
        "trabajo_equipo": "Trabajo en equipo",
        "valores": "Valores personales",
        "recomendacion": "Recomendacion"
    }

    df_excel = df.drop(columns=["id", "candidato_id", "nombre_completo"], errors="ignore")
    df_excel = df_excel.rename(columns=columnas_excel)

    orden_columnas = [
        "Empresa",
        "Nombre del referente",
        "Cargo del referente",
        "Teléfono",
        "Relacion",
        "Desempeno del candidato",
        "Trabajo en equipo",
        "Valores personales",
        "Recomendacion"
    ]

    return df_excel[[col for col in orden_columnas if col in df_excel.columns]]


def crear_excel_referencias(df, candidato, titulo):
    output = BytesIO()
    df_excel = preparar_referencias_para_excel(df)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_excel.to_excel(writer, index=False, sheet_name="Referencias", startrow=5)

        wb = writer.book
        ws = writer.sheets["Referencias"]
        max_col = max(len(df_excel.columns), 1)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)

        ws.cell(row=1, column=1).value = titulo
        ws.cell(row=2, column=1).value = f"Candidato: {candidato}"
        ws.cell(row=3, column=1).value = "Reporte de referencias laborales"

        for row in [1, 2, 3]:
            cell = ws.cell(row=row, column=1)
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(row=1, column=1).font = Font(bold=True, size=16, color=GOLD)
        ws.cell(row=2, column=1).font = Font(bold=True, size=12, color=WHITE)
        ws.cell(row=3, column=1).font = Font(size=11, color=WHITE)

        thin_border = Border(
            left=Side(style="thin", color=BORDER),
            right=Side(style="thin", color=BORDER),
            top=Side(style="thin", color=BORDER),
            bottom=Side(style="thin", color=BORDER)
        )

        header_row = 6

        for cell in ws[header_row]:
            cell.font = Font(bold=True, color=WHITE)
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in ws.iter_rows(min_row=7, max_row=ws.max_row, max_col=max_col):
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for column_cells in ws.iter_cols(min_col=1, max_col=max_col):
            column_letter = get_column_letter(column_cells[0].column)
            max_length = 12

            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[column_letter].width = min(max_length + 4, 38)

        for row_number in range(7, ws.max_row + 1):
            ws.row_dimensions[row_number].height = 45

        ws.freeze_panes = "A7"
        ws.sheet_view.showGridLines = False

    output.seek(0)
    return output.getvalue()


def modulo_referencias():

    conn = sqlite3.connect("seleccion_personal.db", check_same_thread=False)
    cursor = conn.cursor()

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS referencias (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        candidato_id INTEGER,
        empresa TEXT,
        nombre_referente TEXT,
        cargo_referente TEXT,
        telefono TEXT,
        relacion TEXT,
        desempeno TEXT,
        trabajo_equipo TEXT,
        valores TEXT,
        recomendacion TEXT
    )
    """)
    conn.commit()

    st.title("Referencias laborales")

    df_candidatos = pd.read_sql(
        "SELECT id, nombre_completo FROM candidatos ORDER BY nombre_completo",
        conn
    )

    if df_candidatos.empty:
        st.warning("No hay candidatos registrados")
        conn.close()
        return

    candidato = st.selectbox(
        "Seleccione candidato",
        df_candidatos["nombre_completo"]
    )

    candidato_id = df_candidatos[
        df_candidatos["nombre_completo"] == candidato
    ]["id"].values[0]

    with st.form("form_referencias"):

        empresa = st.text_input("Empresa donde trabajó")
        nombre_ref = st.text_input("Nombre del referente")
        cargo_ref = st.text_input("Cargo del referente")
        telefono = st.text_input("Teléfono")
        relacion = st.text_input("Relación (jefe, colega, etc.)")

        st.subheader("Evaluación del referente")

        desempeno = st.text_area("Desempeño del candidato")
        equipo = st.text_area("Trabajo en equipo")
        valores = st.text_area("Valores personales")

        recomendacion = st.selectbox(
            "¿Recomienda su contratación?",
            ["SI", "NO", "CON RESERVAS"]
        )

        guardar = st.form_submit_button("Guardar referencia")

    if guardar:
        cursor.execute("""
            INSERT INTO referencias (
                candidato_id,
                empresa,
                nombre_referente,
                cargo_referente,
                telefono,
                relacion,
                desempeno,
                trabajo_equipo,
                valores,
                recomendacion
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            candidato_id,
            empresa,
            nombre_ref,
            cargo_ref,
            telefono,
            relacion,
            desempeno,
            equipo,
            valores,
            recomendacion
        ))

        conn.commit()
        st.success(" Referencia guardada correctamente")
        st.rerun()

    st.divider()
    st.subheader("Referencias registradas")

    df_ref = pd.read_sql("""
        SELECT
            r.id,
            r.candidato_id,
            c.nombre_completo,
            r.empresa,
            r.nombre_referente,
            r.cargo_referente,
            r.telefono,
            r.relacion,
            r.desempeno,
            r.trabajo_equipo,
            r.valores,
            r.recomendacion
        FROM referencias r
        LEFT JOIN candidatos c ON c.id = r.candidato_id
        WHERE r.candidato_id = ?
        ORDER BY r.id DESC
    """, conn, params=(candidato_id,))

    if df_ref.empty:
        st.info("Este candidato aún no tiene referencias registradas")
    else:
        columnas_vista = [
            "empresa",
            "nombre_referente",
            "cargo_referente",
            "telefono",
            "relacion",
            "recomendacion"
        ]

        st.dataframe(
            df_ref[columnas_vista],
            use_container_width=True,
            hide_index=True
        )

        st.divider()
        st.subheader(" Descargar referencias")

        tipo_descarga = st.radio(
            "Seleccione qué desea descargar",
            [
                "Última referencia registrada",
                "Referencia seleccionada",
                "Todas las referencias del candidato"
            ],
            horizontal=True
        )

        if tipo_descarga == "Última referencia registrada":
            df_descarga = df_ref.head(1)
            titulo_excel = "Última referencia registrada"
            nombre_archivo = f"ultima_referencia_{limpiar_nombre_archivo(candidato)}.xlsx"

        elif tipo_descarga == "Referencia seleccionada":
            opciones = df_ref.apply(
                lambda fila: (
                    f"{fila['id']} - "
                    f"{fila['empresa'] or 'Sin empresa'} - "
                    f"{fila['nombre_referente'] or 'Sin referente'}"
                ),
                axis=1
            )

            seleccion = st.selectbox(
                "Seleccione la referencia que desea descargar",
                opciones
            )

            referencia_id = int(seleccion.split(" - ")[0])
            df_descarga = df_ref[df_ref["id"] == referencia_id]
            titulo_excel = "Referencia laboral seleccionada"
            nombre_archivo = f"referencia_{referencia_id}_{limpiar_nombre_archivo(candidato)}.xlsx"

        else:
            df_descarga = df_ref
            titulo_excel = "Referencias laborales del candidato"
            nombre_archivo = f"referencias_{limpiar_nombre_archivo(candidato)}.xlsx"

        excel = crear_excel_referencias(
            df_descarga,
            candidato=candidato,
            titulo=titulo_excel
        )

        st.download_button(
            label="Descargar Excel profesional",
            data=excel,
            file_name=nombre_archivo,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    conn.close()
