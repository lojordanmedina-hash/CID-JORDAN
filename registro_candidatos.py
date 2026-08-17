import re
import streamlit as st
import sqlite3
import pandas as pd
from io import BytesIO
from datetime import date
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


NAVY = "03045E"
GOLD = "FDD817"
WHITE = "FFFFFF"
LIGHT_BLUE = "EEF3FA"
BORDER = "B7C3D0"


def limpiar_nombre_archivo(nombre):
    nombre_limpio = re.sub(r"[^A-Za-z0-9_-]+", "_", str(nombre).strip())
    return nombre_limpio.strip("_") or "reporte"


def asegurar_columna(cursor, tabla, columna, tipo):
    cursor.execute(f"PRAGMA table_info({tabla})")
    columnas = [fila[1] for fila in cursor.fetchall()]

    if columna not in columnas:
        cursor.execute(f"ALTER TABLE {tabla} ADD COLUMN {columna} {tipo}")


def preparar_candidatos_para_excel(df):
    columnas_excel = {
        "fecha_registro": "Fecha de registro",
        "nombre_completo": "Nombre completo",
        "cargo_postula": "Cargo al que postula",
        "edad": "Edad",
        "estado_civil": "Estado civil",
        "hijos": "Hijos",
        "formacion_tercer_nivel": "Título tercer nivel",
        "titulo_cuarto_nivel": "Título cuarto nivel",
        "anos_experiencia": "Años de experiencia",
        "empresa": "Última empresa",
        "cargo": "Último cargo",
        "actividades": "Actividades realizadas",
        "sueldo_ultimo": "Último sueldo",
        "motivo_salida": "Motivo de salida",
        "disponibilidad": "Disponibilidad",
        "observaciones": "Observaciones"
    }

    df_excel = df.drop(columns=["id"], errors="ignore")
    df_excel = df_excel.rename(columns=columnas_excel)

    orden_columnas = [
        "Fecha de registro",
        "Nombre completo",
        "Cargo al que postula",
        "Edad",
        "Estado civil",
        "Hijos",
        "Título tercer nivel",
        "Título cuarto nivel",
        "Años de experiencia",
        "Última empresa",
        "Último cargo",
        "Actividades realizadas",
        "Último sueldo",
        "Motivo de salida",
        "Disponibilidad",
        "Observaciones"
    ]

    return df_excel[[col for col in orden_columnas if col in df_excel.columns]]


def crear_excel_candidatos(df, titulo):
    output = BytesIO()
    df_excel = preparar_candidatos_para_excel(df)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_excel.to_excel(writer, index=False, sheet_name="Candidatos", startrow=4)

        wb = writer.book
        ws = writer.sheets["Candidatos"]

        max_col = max(len(df_excel.columns), 1)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)

        ws.cell(row=1, column=1).value = titulo
        ws.cell(row=2, column=1).value = "Reporte de selección de personal"

        for row in [1, 2]:
            cell = ws.cell(row=row, column=1)
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(row=1, column=1).font = Font(bold=True, size=16, color=GOLD)
        ws.cell(row=2, column=1).font = Font(bold=True, size=12, color=WHITE)

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 24

        thin_border = Border(
            left=Side(style="thin", color=BORDER),
            right=Side(style="thin", color=BORDER),
            top=Side(style="thin", color=BORDER),
            bottom=Side(style="thin", color=BORDER)
        )

        header_row = 5

        for cell in ws[header_row]:
            cell.font = Font(bold=True, color=WHITE)
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, max_col=max_col):
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for column_cells in ws.iter_cols(min_col=1, max_col=max_col):
            column_letter = get_column_letter(column_cells[0].column)
            max_length = 12

            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[column_letter].width = min(max_length + 4, 40)

        for row_number in range(6, ws.max_row + 1):
            ws.row_dimensions[row_number].height = 38

        ws.freeze_panes = "A6"
        ws.sheet_view.showGridLines = False

    output.seek(0)
    return output.getvalue()


def registro_candidatos():
    conn = sqlite3.connect("seleccion_personal.db", check_same_thread=False)
    cursor = conn.cursor()

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS candidatos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre_completo TEXT,
        edad INTEGER,
        estado_civil TEXT,
        hijos INTEGER,
        formacion_tercer_nivel TEXT,
        titulo_cuarto_nivel TEXT,
        anos_experiencia INTEGER,
        empresa TEXT,
        cargo TEXT,
        actividades TEXT,
        sueldo_ultimo REAL,
        motivo_salida TEXT,
        disponibilidad TEXT,
        observaciones TEXT,
        cargo_postula TEXT,
        fecha_registro TEXT
    )
    """)

    asegurar_columna(cursor, "candidatos", "fecha_registro", "TEXT")
    cursor.execute("""
        UPDATE candidatos
        SET fecha_registro = ?
        WHERE fecha_registro IS NULL OR fecha_registro = ''
    """, (date.today().isoformat(),))

    conn.commit()

    st.title("Registro de candidatos")

    with st.form("form_candidato"):
        col1, col2, col3 = st.columns(3)

        with col1:
            nombre = st.text_input("Nombre completo")
            edad = st.number_input("Edad", min_value=18, max_value=80)
            estado_civil = st.selectbox(
                "Estado civil",
                ["SOLTERO", "CASADO", "UNIÓN LIBRE", "DIVORCIADO"]
            )
            hijos = st.number_input("Hijos", min_value=0)

        with col2:
            cargo_postula = st.text_input("Cargo al que postula")
            titulo_tercer = st.selectbox("Título tercer nivel", ["SI", "NO"])
            titulo_cuarto = st.selectbox("Título cuarto nivel", ["SI", "NO"])
            anos_experiencia = st.number_input("Años de experiencia total", min_value=0)

        with col3:
            fecha_registro = st.date_input("Fecha de registro", value=date.today())
            empresa = st.text_input("Última empresa")
            cargo = st.text_input("Último cargo desempeñado")
            sueldo = st.number_input("Último sueldo", min_value=0.0, step=50.0)
            disponibilidad = st.selectbox(
                "Disponibilidad",
                ["INMEDIATA", "15 DÍAS", "30 DÍAS"]
            )

        actividades = st.text_area("Actividades realizadas")
        motivo = st.text_area("Motivo de salida")
        observaciones = st.text_area("Observaciones")

        guardar = st.form_submit_button("Guardar candidato")

    if guardar:
        if not nombre.strip():
            st.warning("Debe ingresar el nombre completo del candidato")
        elif not cargo_postula.strip():
            st.warning("Debe ingresar el cargo al que postula")
        else:
            cursor.execute("""
                INSERT INTO candidatos (
                    nombre_completo,
                    edad,
                    estado_civil,
                    hijos,
                    formacion_tercer_nivel,
                    titulo_cuarto_nivel,
                    anos_experiencia,
                    empresa,
                    cargo,
                    actividades,
                    sueldo_ultimo,
                    motivo_salida,
                    disponibilidad,
                    observaciones,
                    cargo_postula,
                    fecha_registro
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                nombre,
                edad,
                estado_civil,
                hijos,
                titulo_tercer,
                titulo_cuarto,
                anos_experiencia,
                empresa,
                cargo,
                actividades,
                sueldo,
                motivo,
                disponibilidad,
                observaciones,
                cargo_postula,
                fecha_registro.isoformat()
            ))

            conn.commit()
            st.success(" Candidato guardado correctamente")
            st.rerun()

    st.divider()
    st.subheader("Candidatos registrados")

    df = pd.read_sql("""
        SELECT *
        FROM candidatos
        ORDER BY id DESC
    """, conn)

    if df.empty:
        st.info("Aún no hay candidatos registrados")
        conn.close()
        return

    df["fecha_registro"] = pd.to_datetime(df["fecha_registro"], errors="coerce")

    st.subheader("Filtros")

    col_f1, col_f2, col_f3 = st.columns(3)

    with col_f1:
        cargos = sorted(df["cargo_postula"].dropna().unique().tolist())
        cargo_filtro = st.selectbox(
            "Filtrar por cargo",
            ["TODOS"] + cargos
        )

    with col_f2:
        fecha_inicio = st.date_input(
            "Desde",
            value=df["fecha_registro"].min().date()
        )

    with col_f3:
        fecha_fin = st.date_input(
            "Hasta",
            value=df["fecha_registro"].max().date()
        )

    df_filtrado = df.copy()

    if cargo_filtro != "TODOS":
        df_filtrado = df_filtrado[df_filtrado["cargo_postula"] == cargo_filtro]

    df_filtrado = df_filtrado[
        (df_filtrado["fecha_registro"].dt.date >= fecha_inicio) &
        (df_filtrado["fecha_registro"].dt.date <= fecha_fin)
    ]

    columnas_vista = [
        "fecha_registro",
        "nombre_completo",
        "cargo_postula",
        "edad",
        "estado_civil",
        "anos_experiencia",
        "empresa",
        "cargo",
        "sueldo_ultimo",
        "disponibilidad"
    ]

    df_vista = df_filtrado[columnas_vista].copy()
    df_vista["fecha_registro"] = df_vista["fecha_registro"].dt.strftime("%Y-%m-%d")

    st.dataframe(
        df_vista,
        use_container_width=True,
        hide_index=True
    )

    st.divider()
    st.subheader("Descargar candidatos")

    tipo_descarga = st.radio(
        "Seleccione qué desea descargar",
        [
            "Últimos 3 candidatos registrados",
            "Candidatos filtrados por cargo y fecha",
            "Todos los candidatos"
        ],
        horizontal=True
    )

    if tipo_descarga == "Últimos 3 candidatos registrados":
        df_descarga = df.sort_values("id", ascending=False).head(3)
        titulo_excel = "Últimos 3 candidatos registrados"
        nombre_archivo = "ultimos_3_candidatos.xlsx"

    elif tipo_descarga == "Candidatos filtrados por cargo y fecha":
        df_descarga = df_filtrado
        titulo_excel = "Candidatos filtrados"
        nombre_archivo = f"candidatos_filtrados_{limpiar_nombre_archivo(cargo_filtro)}.xlsx"

    else:
        df_descarga = df
        titulo_excel = "Todos los candidatos registrados"
        nombre_archivo = "todos_los_candidatos.xlsx"

    df_descarga = df_descarga.copy()
    df_descarga["fecha_registro"] = df_descarga["fecha_registro"].dt.strftime("%Y-%m-%d")

    excel = crear_excel_candidatos(df_descarga, titulo_excel)

    st.download_button(
        label="Descargar Excel profesional",
        data=excel,
        file_name=nombre_archivo,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    conn.close()
